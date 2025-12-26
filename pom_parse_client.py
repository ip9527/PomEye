import ttkbootstrap as ttk
import tkinter
import tkinter.filedialog
import tkinter.messagebox
from tkinter import Canvas, StringVar, END, TOP, X, RIGHT, VERTICAL
import webbrowser
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from variables import *
from ttkbootstrap.constants import *
from parse import parse
from check_vul import get_details_by_version
from config import (
    THEME_NAME,
    TABLE_HEIGHT,
    MAIN_WINDOW_WIDTH_RATIO,
    MAIN_WINDOW_HEIGHT_RATIO,
    RESULT_WINDOW_FULLSCREEN,
    LEVEL_SORT_ORDER,
    LEVEL_COLORS,
    LEVEL_MAPPING
)

'''
可视化界面模块：上传界面，详细信息界面，跳转
'''


# 上传文件的界面，也就是主界面
def upload_gui():
    root = ttk.Window("pom文件提取组件版本及漏洞检测", themename=THEME_NAME)
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.geometry(f'{int(MAIN_WINDOW_WIDTH_RATIO * screen_width)}x{int(MAIN_WINDOW_HEIGHT_RATIO * screen_height)}')

    global progressbar_tips
    progressbar_tips = StringVar(root)
    progressbar_tips.set("还未上传文件！！！")

    # 说明文件
    t1 = ttk.Label(root, text="上传你需要检测的pom文件：\n    1.上传所有需要检测的pom文件\n    2.上传文件夹，程序会自动查找pom文件\n注：程序会解析父子项目依赖关系",
                   bootstyle="danger")
    t1.pack(anchor="center", expand=True)

    # 进度条
    p1 = ttk.Progressbar(root, bootstyle="info-striped", length=int(1 * (screen_width / 5)))

    upload_frame = ttk.Frame(root)
    upload_frame.pack(anchor="center", expand=True)
    # 上传按钮
    b1 = ttk.Button(upload_frame, text="上传文件", bootstyle=(INFO, OUTLINE),
                    command=lambda: upload_to_info_gui(root, p1, "openFiles"))
    b1.pack(side='left', padx=20)
    b2 = ttk.Button(upload_frame, text="上传文件夹", bootstyle=(SUCCESS, OUTLINE),
                    command=lambda: upload_to_info_gui(root, p1, "openDir"))
    b2.pack(side='left', padx=20)

    # 进度条文字
    t2 = ttk.Label(root, textvariable=progressbar_tips, bootstyle="info")
    t2.pack()
    p1.pack(anchor="center", side=BOTTOM, pady=40)

    root.mainloop()


# 文件上传后跳转到组件详细信息界面
def upload_to_info_gui(root, p1, type):
    if type == "openFiles":
        files = tkinter.filedialog.askopenfilenames()
    else:
        files = tkinter.filedialog.askdirectory()
    
    # 检查用户是否取消了文件选择
    if not files:
        progressbar_tips.set("已取消文件选择")
        return
    
    # 解析文件并检测漏洞（阻塞直到完成）
    parse(files, p1, root, progressbar_tips)
    
    # 检查是否解析到数据
    if not xml_res:
        progressbar_tips.set("⚠️ 警告：未解析到任何组件数据，请检查文件格式")
        return
    
    # 显示完成提示
    progressbar_tips.set("✓ 检测完成！正在打开结果页面...")
    root.update()
    # 检测完成后才打开结果窗口
    info_gui(root)


# 组件详细信息界面
def info_gui(root):
    # 检查数据
    if not xml_res:
        # 显示错误提示窗口
        error_window = ttk.Toplevel(root)
        error_window.title("错误")
        error_window.geometry("400x150")
        error_label = ttk.Label(error_window, text="❌ 未找到任何组件数据\n\n请检查：\n1. 上传的文件是否为有效的 pom.xml\n2. 文件中是否包含 <dependency> 标签", 
                               bootstyle="danger", font=('Arial', 12))
        error_label.pack(expand=True)
        return
    
    # 主窗口
    root2 = ttk.Toplevel(root)
    root2.title("pom文件提取组件版本及漏洞检测")
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    # 根据配置决定是否全屏
    if RESULT_WINDOW_FULLSCREEN:
        root2.geometry(f'{int(screen_width)}x{int(screen_height)}')
    else:
        root2.geometry(f'{int(screen_width / 2)}x{int(screen_height / 2)}')

    # 顶栏提示信息
    lbl = ttk.Label(master=root2, text=f"点击组件查看漏洞详情（共 {len(xml_res)} 个组件）", bootstyle=(LIGHT, INVERSE), anchor="center")
    lbl.pack(side=TOP, fill=X)
    
    # 导出按钮框架
    export_frame = ttk.Frame(root2)
    export_frame.pack(side=TOP, fill=X, pady=5)
    
    def export_to_excel():
        """导出检测结果到 Excel 文件"""
        if not xml_res:
            return
        
        # 选择保存路径
        file_path = tkinter.filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            title="保存检测结果"
        )
        
        if not file_path:
            return
        
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "漏洞检测结果"
            
            # 定义样式
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 定义漏洞等级颜色
            level_colors = {
                '严重': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
                '高危': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
                '中危': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
                '低危': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'),
                '请求失败': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
            }
            
            # 设置表头
            headers = ['序号', 'groupId', 'artifactId', 'version', '漏洞等级', '漏洞名称', 'CVE编号', 'CWE编号', '影响版本范围', '漏洞概述', '详情链接', '来源文件']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 设置列宽
            column_widths = [6, 30, 30, 12, 10, 40, 15, 15, 20, 60, 50, 40]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = width
            
            # 按漏洞等级排序后的数据
            sorted_data = sorted(xml_res, key=lambda x: LEVEL_SORT_ORDER.get(x[3], 6))
            
            # 填充数据
            row_num = 2
            for idx, info in enumerate(sorted_data, 1):
                ga = f"{info[0]}:{info[1]}"
                version = info[2]
                level = info[3]
                source_file = info[4]
                
                # 获取漏洞详情
                vul_details_list = get_details_by_version(ga, version)
                
                if vul_details_list and len(vul_details_list) > 0:
                    # 有漏洞，每个漏洞一行
                    for vul in vul_details_list:
                        ws.cell(row=row_num, column=1, value=idx)
                        ws.cell(row=row_num, column=2, value=info[0])
                        ws.cell(row=row_num, column=3, value=info[1])
                        ws.cell(row=row_num, column=4, value=version)
                        ws.cell(row=row_num, column=5, value=level)
                        ws.cell(row=row_num, column=6, value=vul.name)
                        ws.cell(row=row_num, column=7, value=vul.cve)
                        ws.cell(row=row_num, column=8, value=vul.cwe)
                        ws.cell(row=row_num, column=9, value=f"[{vul.min_version}, {vul.max_version})")
                        ws.cell(row=row_num, column=10, value=vul.overview)
                        ws.cell(row=row_num, column=11, value=vul.href)
                        ws.cell(row=row_num, column=12, value=source_file)
                        
                        # 设置漏洞等级背景色
                        if level in level_colors:
                            for col in range(1, 13):
                                ws.cell(row=row_num, column=col).fill = level_colors[level]
                        
                        row_num += 1
                else:
                    # 无漏洞或请求失败
                    ws.cell(row=row_num, column=1, value=idx)
                    ws.cell(row=row_num, column=2, value=info[0])
                    ws.cell(row=row_num, column=3, value=info[1])
                    ws.cell(row=row_num, column=4, value=version)
                    ws.cell(row=row_num, column=5, value=level)
                    ws.cell(row=row_num, column=6, value='-')
                    ws.cell(row=row_num, column=7, value='-')
                    ws.cell(row=row_num, column=8, value='-')
                    ws.cell(row=row_num, column=9, value='-')
                    ws.cell(row=row_num, column=10, value='无漏洞' if level == '*' else '请求失败')
                    ws.cell(row=row_num, column=11, value='-')
                    ws.cell(row=row_num, column=12, value=source_file)
                    
                    # 设置漏洞等级背景色
                    if level in level_colors:
                        for col in range(1, 13):
                            ws.cell(row=row_num, column=col).fill = level_colors[level]
                    elif level == '*':
                        # 无漏洞使用绿色
                        no_vul_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        for col in range(1, 13):
                            ws.cell(row=row_num, column=col).fill = no_vul_fill
                    
                    row_num += 1
            
            # 保存文件
            wb.save(file_path)
            tkinter.messagebox.showinfo("成功", f"检测结果已成功导出到:\n{file_path}")
            
        except Exception as e:
            tkinter.messagebox.showerror("错误", f"导出 Excel 文件时出错:\n{str(e)}")
    
    # 导出按钮
    export_btn = ttk.Button(export_frame, text="📥 导出 Excel", bootstyle=(SUCCESS, OUTLINE),
                            command=export_to_excel)
    export_btn.pack(side=LEFT, padx=10)
    
    # 信息表格
    table_frame = ttk.Frame(root2)
    table_frame.pack(fill=X)
    columns = ["groupId", "artifactId", "version", "level", "fold"]
    
    # 创建自定义样式，设置标题栏为淡紫色
    style = ttk.Style()
    
    # 配置 Treeview 的基础样式
    style.configure('Treeview',
                   rowheight=28,  # 行高
                   font=('Arial', 10))  # 字体
    
    # 配置 Treeview.Heading 样式 (标题栏)
    # 注意：ttkbootstrap 可能需要使用 theme_use 前再配置
    style.configure('Treeview.Heading',
                   background='#E6D5FF',  # 淡紫色背景
                   foreground='#333333',  # 深灰色文字
                   font=('Arial', 10, 'bold'),  # 加粗字体
                   relief='raised',  # 突起效果
                   borderwidth=2)
    style.map('Treeview.Heading',
             background=[('!active', '#E6D5FF'), ('active', '#D1C4E9')],
             relief=[('active', 'groove'), ('!active', 'raised')])
    
    table = ttk.Treeview(
        master=table_frame,  # 父容器
        height=TABLE_HEIGHT,  # 高度,可显示 height 行（从配置文件读取）
        columns=columns,  # 显示的列
        show='headings',
    )
    #滚动条
    scrollbar_y = ttk.Scrollbar(table_frame, orient=VERTICAL, bootstyle="info")
    scrollbar_y.config(command=table.yview)
    table.configure(yscrollcommand=scrollbar_y.set)
    scrollbar_y.pack(side=RIGHT, fill="y")

    for c in columns:
        table.heading(c, text=c)
        table.column(c, anchor='center', width=int(screen_width / 10), stretch=True)
    
    # 按漏洞等级排序：严重 > 高危 > 中危 > 低危 > 无 > 请求失败（从配置文件读取）
    sorted_xml_res = sorted(xml_res, key=lambda x: LEVEL_SORT_ORDER.get(x[3], 6))
    
    # 给表格添加元素，根据漏洞危险程度标记上颜色（使用配置文件）
    for info in sorted_xml_res:
        level = info[3]
        if level in LEVEL_COLORS:
            color_config = LEVEL_COLORS[level]
            if isinstance(color_config, tuple):
                # (背景色, 文字色)
                bg_color, fg_color = color_config
                tag_name = f'level_{level}'
                table.insert("", END, values=info, tags=(tag_name,))
                table.tag_configure(tag_name, background=bg_color, foreground=fg_color)
            else:
                # 只有背景色
                tag_name = f'level_{level}'
                table.insert("", END, values=info, tags=(tag_name,))
                table.tag_configure(tag_name, background=color_config)
        else:
            table.insert("", END, values=info)
    
    # 定义每个等级选中时的加深颜色
    level_colors_selected = {
        '严重': ('#8B0000', '#FFFFFF'),      # 暗红色 + 白色文字
        '高危': ('#CC6600', '#FFFFFF'),      # 暗橙色 + 白色文字
        '中危': ('#B8860B', '#FFFFFF'),      # 暗金黄色 + 白色文字
        '低危': ('#4682B4', '#FFFFFF'),      # 暗蓝色 + 白色文字
        '请求失败': ('#696969', '#FFFFFF'),  # 暗灰色 + 白色文字
    }
    
    # 用于跟踪当前选中的项
    current_selected_item = [None]  # 使用列表以便在闭包中修改

    def treeviewClick(event):  # 单击
        # 获取点击的项（即使还未被选中）
        region = table.identify_region(event.x, event.y)
        if region != 'cell':
            return
        
        clicked_item = table.identify_row(event.y)
        if not clicked_item:
            return
        
        item_text = table.item(clicked_item, "values")
        
        # 恢复之前选中项的颜色
        if current_selected_item[0]:
            item_values = table.item(current_selected_item[0], "values")
            level = item_values[3]
            if level in LEVEL_COLORS:
                color_config = LEVEL_COLORS[level]
                if isinstance(color_config, tuple):
                    bg_color, fg_color = color_config
                    table.item(current_selected_item[0], tags=(f'level_{level}',))
                else:
                    table.item(current_selected_item[0], tags=(f'level_{level}',))
            else:
                table.item(current_selected_item[0], tags=())
        
        # 为点击的行设置加深的颜色
        selected_level = item_text[3]
        if selected_level in level_colors_selected:
            selected_bg, selected_fg = level_colors_selected[selected_level]
            selected_tag = f'selected_{selected_level}'
            table.tag_configure(selected_tag, background=selected_bg, foreground=selected_fg)
            table.item(clicked_item, tags=(selected_tag,))
        elif selected_level == '*':  # 处理无漏洞的情况
            selected_tag = 'selected_no_vul'
            table.tag_configure(selected_tag, background='#2E8B57', foreground='#FFFFFF')
            table.item(clicked_item, tags=(selected_tag,))
        
        # 更新当前选中项
        current_selected_item[0] = clicked_item
        
        # 取消 Treeview 的默认选中样式（重要！）
        table.selection_remove(table.selection())
        print(item_text)
        try:
            text.delete("1.0", END)
        except:
            pass
        #在下方的文本框显示漏洞详情
        vul_details = get_details_by_version(item_text[0]+":"+item_text[1],item_text[2])
        # text_content = ""
        # for d in vul_details:
        #     text_content = text_content +f"{d.name}\n{d.cve}\n"+"="*10+"\n"
        # text.insert(INSERT, text_content)
        # 因为设置文本框很繁琐，我就放到另一个方法里了
        info_text_gui(text,vul_details,item_text[0]+":"+item_text[1],screen_width)

    # 给表格绑定点击事件
    table.bind('<ButtonRelease-1>', treeviewClick)
    table.pack(fill=X)

    # 详情栏
    text = ttk.Text(root2, undo=True, autoseparators=False)
    
    # 创建一个 Tooltip 标签（初始隐藏）
    tooltip = ttk.Label(root2, text="", bootstyle="info", font=('Arial', 10), 
                        relief="solid", borderwidth=1, padding=5)
    tooltip.place_forget()  # 初始隐藏
    
    # 为文本框绑定双击和鼠标事件（只绑定一次）
    def on_double_click(event):
        # 获取双击位置的文本
        index = text.index(f"@{event.x},{event.y}")
        # 查找该位置的标签
        tags = text.tag_names(index)
        if 'url_tag' in tags:
            # 获取整行 URL
            line_start = text.index(f"{index} linestart")
            line_end = text.index(f"{index} lineend")
            url_text = text.get(line_start, line_end).strip()
            # 提取 URL（https://...）
            url_match = re.search(r'https?://[^\s]+', url_text)
            if url_match:
                url = url_match.group(0)
                webbrowser.open(url)
    
    def on_enter(event):
        # 鼠标进入 URL 时改变样式并显示提示
        index = text.index(f"@{event.x},{event.y}")
        tags = text.tag_names(index)
        if 'url_tag' in tags:
            text.config(cursor="hand2")  # 改成手形光标
            # 显示工具提示
            tooltip_text = "👆 双击跳转到详情页面"
            tooltip.config(text=tooltip_text)
            # 计算提示位置（鼠标下方）
            x = event.x_root - root2.winfo_rootx() + 10
            y = event.y_root - root2.winfo_rooty() + 20
            tooltip.place(x=x, y=y)
        else:
            # 不在 URL 上时隐藏提示
            tooltip.place_forget()
    
    def on_leave(event):
        # 鼠标离开时恢复光标并隐藏提示
        text.config(cursor="")
        tooltip.place_forget()
    
    def on_motion(event):
        # 鼠标移动时实时检查是否在 URL 上
        index = text.index(f"@{event.x},{event.y}")
        tags = text.tag_names(index)
        if 'url_tag' in tags:
            text.config(cursor="hand2")
            # 更新提示位置
            tooltip_text = "👆 双击跳转到详情页面"
            tooltip.config(text=tooltip_text)
            x = event.x_root - root2.winfo_rootx() + 10
            y = event.y_root - root2.winfo_rooty() + 20
            tooltip.place(x=x, y=y)
        else:
            text.config(cursor="")
            tooltip.place_forget()
    
    text.bind("<Double-Button-1>", on_double_click)
    text.bind("<Motion>", on_motion)
    text.bind("<Leave>", on_leave)

    text.pack(side=BOTTOM, fill=X)

    root2.mainloop()

def info_text_gui(text,vul_details,ga,screen_width):
    class TextSeparat(Canvas):  # working
        '''
        用于在tkinter文本框插入不同颜色、样式的分割线
        '''

        def __init__(self, text, width, bg='white', color='#66CCCC', line='common'):
            super().__init__(text, width=width, height=8, background=bg, highlightthickness=0, relief='flat', bd=0)
            if line == 'common':  # ---
                self.create_line(0, 4, width, 4, fill=color, width=2)
            elif line == 'dash':  # - -
                self.create_line(0, 4, width, 4, fill=color, dash=(10, 3), width=2)
            elif line == 'dash_point':  # -··
                self.create_line(0, 4, width, 4, fill=color, dash=(5, 2, 3), width=2)
            elif line == 'point':  # ···
                self.create_line(0, 4, width, 4, fill=color, dash=(2, 2), width=2)
            elif line == 'double_line':  # ===
                self.create_line(0, 3, width, 3, fill=color, width=1)
                self.create_line(0, 6, width, 6, fill=color, width=1)
            elif line == 'double_dash':  # = =
                self.create_line(0, 3, width, 3, fill=color, dash=(10, 3), width=1)
                self.create_line(0, 6, width, 6, fill=color, dash=(10, 3), width=1)

    # 定义标签样式
    font1 = ('Arial', 16, 'bold')
    text.tag_configure('bold_style', font=font1)
    font2 = ('Arial', 13)
    text.tag_configure('not_bold_style', font=font2)
    text.tag_configure('red_style', font=font2, foreground='#FF4500')
    text.tag_configure('blue_style', font=font2, foreground='#6495ED')
    text.tag_configure('deep_blue_style', font=font2, foreground='#7B68EE')
    text.tag_configure('gray_style', font=font2, foreground='#696969')
    # URL 标签样式：蓝色、下划线、可点击
    text.tag_configure('url_tag', font=font2, foreground='#6495ED', underline=True)

    text.window_create('end', window=TextSeparat(text, screen_width, bg=text['background'], line='double_line'))
    text.insert(END, "\n\n")
    for v in vul_details:
        text.insert(INSERT, f"{v.name}", 'bold_style')
        # 使用配置文件中的漏洞等级映射
        level = "*"
        for level_code, level_name in LEVEL_MAPPING.items():
            if level_code in v.level:
                level = level_name
                break
        text.insert(INSERT, f"      {level}", 'red_style')
        text.insert(INSERT, "\nAffecting ", 'not_bold_style')
        text.insert(INSERT, f"{ga}", 'red_style')
        text.insert(INSERT, " package, versions ", 'not_bold_style')
        text.insert(INSERT, f"[{v.min_version}, {v.max_version})", 'red_style')
        text.insert(INSERT, f"\n{v.cve}      {v.cwe}", 'deep_blue_style')
        text.insert(INSERT,
                f"\n{v.overview}",
                'gray_style')
        text.insert(INSERT, f"\n{v.href}", 'url_tag')

        text.insert(END, "\n\n")
        text.window_create('end', window=TextSeparat(text, screen_width, bg=text['background'], line='double_line'))
        text.insert(END, "\n\n")
    # 将文字设置为居中
    text.tag_add("center", "1.0", "end")
    text.tag_configure("center", justify="center")
